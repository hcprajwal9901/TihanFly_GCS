import os
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Configuration
ROOT_DIR = r'E:\TiHAN\TflyGCS (1)\TflyGCS'
OUT_FILE = r'E:\TiHAN\TflyGCS (1)\TflyGCS\frontend_module_function_descrption_for_unit_testing.xlsx'

# Exclude directories
EXCLUDE_DIRS = ['node_modules', 'build', 'dist', 'TihanFlyCC-main', 'leaflet', '.git']
TARGET_EXTS = ['.js', '.ts', '.html']

def get_latest_commit_id(repo_dir):
    try:
        result = subprocess.run(
            ['git', 'rev-parse', 'HEAD'],
            cwd=repo_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )
        return result.stdout.strip()
    except Exception as e:
        print(f"Error getting commit ID: {e}")
        return "Unknown"

def generate_description(func_name):
    name_clean = func_name.replace('_', ' ')
    name_clean = re.sub('([a-z])([A-Z])', r'\g<1> \g<2>', name_clean).lower()
    if name_clean.startswith('get'):
        return f"Retrieves the {name_clean[4:]}."
    elif name_clean.startswith('set'):
        return f"Sets the {name_clean[4:]}."
    elif name_clean.startswith('is') or name_clean.startswith('has') or name_clean.startswith('can'):
        return f"Checks if it {name_clean}."
    elif name_clean.startswith('handle') or name_clean.startswith('on'):
        return f"Event handler for {name_clean}."
    else:
        return f"Handles functionality for {name_clean}."

def get_files(root_dir):
    matched_files = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        # normalize dirpath to compare with EXCLUDE_DIRS
        norm_dirpath = dirpath.replace('\\', '/')
        if any(excl in norm_dirpath for excl in EXCLUDE_DIRS):
            continue
        for f in filenames:
            if any(f.endswith(ext) for ext in TARGET_EXTS):
                matched_files.append(os.path.join(dirpath, f))
    return matched_files

def extract_functions(filepath):
    functions = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            lines = content.split('\n')
            
            p1 = re.compile(r'(?:async\s+)?function\s+([a-zA-Z0-9_]+)\s*\(([^)]*)\)')
            p2 = re.compile(r'(?:const|let|var)\s+([a-zA-Z0-9_]+)\s*=\s*(?:async\s+)?(?:\([^)]*\)|[a-zA-Z0-9_]+)\s*=>')
            p3 = re.compile(r'^\s*(?:async\s+)?([a-zA-Z0-9_]+)\s*\(([^)]*)\)\s*\{')

            for i, line in enumerate(lines):
                line = line.strip()
                if not line or line.startswith('//') or line.startswith('/*') or line.startswith('*'):
                    continue
                
                func_name = None
                args = 'None'
                ret_type = 'Dynamic / Any'
                
                m1 = p1.search(line)
                m2 = p2.search(line)
                m3 = p3.search(line)
                
                if m1:
                    func_name = m1.group(1)
                    args = m1.group(2)
                    if 'async' in line:
                        ret_type = 'Promise (Async)'
                elif m2:
                    func_name = m2.group(1)
                    if 'async' in line:
                        ret_type = 'Promise (Async)'
                    args_match = re.search(r'\(([^)]*)\)\s*=>', line)
                    if args_match:
                        args = args_match.group(1)
                    else:
                        args_match2 = re.search(r'=\s*([a-zA-Z0-9_]+)\s*=>', line)
                        if args_match2:
                            args = args_match2.group(1)
                elif m3:
                    func_name = m3.group(1)
                    args = m3.group(2)
                    if 'async' in line:
                        ret_type = 'Promise (Async)'
                    if func_name in ['if', 'for', 'while', 'catch', 'switch', 'return', 'else', 'function']:
                        func_name = None
                        
                if func_name:
                    if ret_type == 'Dynamic / Any':
                        if func_name[0].isupper():
                            ret_type = 'UI Component / Object'
                        elif func_name.startswith('on') or func_name.startswith('handle') or func_name.startswith('set'):
                            ret_type = 'Executes side-effects (Void)'
                        elif func_name.startswith('is') or func_name.startswith('has') or func_name.startswith('can'):
                            ret_type = 'Boolean'
                        else:
                            ret_type = 'Dynamic (Check implementation)'

                        
                if func_name:
                    functions.append({
                        'name': func_name,
                        'line': i + 1,
                        'input': args if args and args.strip() else 'None',
                        'expected_output': ret_type,
                        'description': generate_description(func_name)
                    })
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
    return functions

def load_existing_functions(filepath):
    existing_funcs = set()
    if not os.path.exists(filepath):
        return existing_funcs
    
    try:
        wb = openpyxl.load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            current_file = None
            for row in ws.iter_rows(min_row=1, values_only=True):
                for cell_val in row:
                    if isinstance(cell_val, str) and cell_val.startswith('File: '):
                        current_file = cell_val[6:].strip()
                        break
                    elif isinstance(cell_val, str) and current_file and ' : ' in cell_val:
                        func_name = cell_val.split(' : ')[0].strip()
                        existing_funcs.add((current_file, func_name))
                        break
    except Exception as e:
        print(f"Error loading existing workbook: {e}")
        
    return existing_funcs

def main():
    print("Scanning frontend codebase for functions...")
    files = get_files(ROOT_DIR)
    
    print("Loading previously recorded functions from Excel...")
    existing_funcs = load_existing_functions(OUT_FILE)
    
    new_or_updated_entries = {}
    
    for filepath in files:
        rel_path = os.path.relpath(filepath, ROOT_DIR)
        functions = extract_functions(filepath)
        
        for func in functions:
            func_key = (rel_path, func['name'])
            if func_key not in existing_funcs:
                if rel_path not in new_or_updated_entries:
                    new_or_updated_entries[rel_path] = []
                new_or_updated_entries[rel_path].append(func)

    if not new_or_updated_entries:
        print("No new changes or functions detected. The Excel sheet is up to date.")
        return

    print(f"Detected {sum(len(funcs) for funcs in new_or_updated_entries.values())} new/modified functions.")

    commit_id = get_latest_commit_id(ROOT_DIR)

    # Load the workbook to append a new sheet
    if os.path.exists(OUT_FILE):
        wb = openpyxl.load_workbook(OUT_FILE)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        sheet_name = f"Changes_{timestamp}"
        ws = wb.create_sheet(title=sheet_name)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        sheet_name = "Unit Testing Descriptions"
        ws = wb.create_sheet(title=sheet_name)
        
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    file_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font_bold = Font(bold=True)
    
    current_row = 1

    # Add commit ID at the top of the file/sheet
    ws.cell(row=current_row, column=1, value=f"Commit ID: {commit_id}").font = font_bold
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 2
    
    for rel_path, functions in new_or_updated_entries.items():
        # Add file name row
        ws.cell(row=current_row, column=1, value=f"File: {rel_path}").font = font_bold
        ws.cell(row=current_row, column=1).fill = file_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 1
        
        # Add headers
        headers = ["Sl No.", "Module/Function (Name : Line Number)", "Description", "Actual Input", "Expected Output"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = font_bold
            cell.fill = header_fill
        current_row += 1
        
        # Add data
        sl_no = 1
        for func in functions:
            name_line = f"{func['name']} : {func['line']}"
            ws.cell(row=current_row, column=1, value=sl_no)
            ws.cell(row=current_row, column=2, value=name_line)
            ws.cell(row=current_row, column=3, value=func['description'])
            ws.cell(row=current_row, column=4, value=func['input'])
            ws.cell(row=current_row, column=5, value=func['expected_output'])
            current_row += 1
            sl_no += 1
            
        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40

    wb.save(OUT_FILE)
    print(f"Successfully generated/added sheet '{sheet_name}' to {OUT_FILE}")

if __name__ == '__main__':
    main()
