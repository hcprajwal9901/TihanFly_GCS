import os
import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Configuration
ROOT_DIR = r'E:\TiHAN\TflyGCS (1)\TflyGCS\TihanFlyCC-main'
OUT_FILE = r'E:\TiHAN\TflyGCS (1)\TflyGCS\backend_module_function_description_for_unit_testing.xlsx'

# Active subdirectories in TihanFlyCC-main to scan
ACTIVE_DIRS = ['Command', 'Firmware', 'Flightmode', 'Inspector', 'Link', 'Parameters', 'Parser', 'Transport', 'Vehicle', 'calibration']
TARGET_EXTS = ['.h', '.hpp', '.cpp']

# Reserved keywords to filter out false positives
KEYWORDS = {'if', 'for', 'while', 'catch', 'switch', 'return', 'else', 'using', 'namespace',
            'class', 'struct', 'typedef', 'template', 'static_assert', 'void', 'const', 'noexcept'}

def get_latest_commit_id(repo_dir):
    try:
        result = subprocess.run(
            ['git', 'rev-parse', 'HEAD'],
            cwd=repo_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True
        )
        return result.stdout.strip()
    except Exception as e:
        print(f"Error getting commit ID: {e}")
        return "Unknown"

def generate_description(func_name):
    # Extract only the base function name (strip class namespace)
    name_clean = func_name.split('::')[-1]
    # Remove destructor tilde if present
    name_clean = name_clean.replace('~', 'destruct ')
    name_clean = name_clean.replace('_', ' ')
    name_clean = re.sub('([a-z])([A-Z])', r'\g<1> \g<2>', name_clean).lower()
    
    if name_clean.startswith('get'):
        return f"Retrieves the {name_clean[4:]}."
    elif name_clean.startswith('set'):
        return f"Sets the {name_clean[4:]}."
    elif name_clean.startswith('is') or name_clean.startswith('has') or name_clean.startswith('check'):
        return f"Checks or validates {name_clean}."
    elif name_clean.startswith('test'):
        return f"Runs unit test for {name_clean[5:]}."
    elif name_clean.startswith('destruct'):
        return f"Destructor for cleaning up resources of the object."
    else:
        return f"Handles functionality for {name_clean}."

def get_files(root_dir):
    matched_files = []
    for active_dir in ACTIVE_DIRS:
        dir_path = os.path.join(root_dir, active_dir)
        if not os.path.exists(dir_path):
            continue
        for dirpath, _, filenames in os.walk(dir_path):
            for f in filenames:
                if any(f.endswith(ext) for ext in TARGET_EXTS):
                    matched_files.append(os.path.join(dirpath, f))
    return matched_files

def extract_functions(filepath):
    functions = []
    
    # Robust patterns for C++ function declaration/definition
    func_pattern = re.compile(
        r'^\s*(?:inline|static|virtual|explicit|friend|const)?\s*'
        r'([a-zA-Z_][a-zA-Z0-9_<>:*\s&]*)\s+'  # Return type
        r'([a-zA-Z_][a-zA-Z0-9_:]*)\s*'        # Function/Method name
        r'\(([^)]*)\)'                         # Arguments
        r'(?:\s*const)?(?:\s*override|\s*final)?\s*(?:[;{]|\bnoexcept\b)?'
    )
    
    # Pattern for constructor / destructor
    ctor_pattern = re.compile(
        r'^\s*(?:explicit)?\s*(~?[a-zA-Z_][a-zA-Z0-9_]*::~?[a-zA-Z_][a-zA-Z0-9_]*|~?[a-zA-Z_][a-zA-Z0-9_]*)\s*\(([^)]*)\)\s*(?:[;{]|\bnoexcept\b)?'
    )

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            lines = content.split('\n')
            
            for i, line in enumerate(lines):
                line = line.strip()
                # Skip comments and preprocessor directives
                if not line or line.startswith('//') or line.startswith('#') or line.startswith('/*') or line.startswith('*'):
                    continue
                
                # Try regular function match
                m = func_pattern.match(line)
                if m:
                    ret_type = m.group(1).strip()
                    func_name = m.group(2).strip()
                    args = m.group(3).strip()
                    
                    # Filter out assignments or false keyword matches
                    if '=' in line and not ('=' in args or '=' in ret_type):
                        # likely an assignment statement if = is outside args and ret_type
                        continue
                    
                    if ret_type in KEYWORDS or func_name in KEYWORDS:
                        continue
                    if any(kw in func_name for kw in ['TEST', 'TEST_F', 'EXPECT_', 'ASSERT_']):
                        continue
                    
                    functions.append({
                        'name': func_name,
                        'line': i + 1,
                        'input': args if args else 'None',
                        'expected_output': ret_type if ret_type and ret_type != 'void' else 'Executes successfully without return',
                        'description': generate_description(func_name)
                    })
                    continue

                # Try constructor/destructor match
                m_ctor = ctor_pattern.match(line)
                if m_ctor:
                    func_name = m_ctor.group(1).strip()
                    args = m_ctor.group(2).strip()
                    
                    # Ensure it's not a control flow keyword
                    if func_name in KEYWORDS or func_name in ['if', 'for', 'while', 'switch', 'catch']:
                        continue
                        
                    functions.append({
                        'name': func_name,
                        'line': i + 1,
                        'input': args if args else 'None',
                        'expected_output': 'Instantiates / destroys object',
                        'description': generate_description(func_name)
                    })
                    
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        
    return functions

def main():
    print("Scanning active backend codebase for functions...")
    files = get_files(ROOT_DIR)
    print(f"Found {len(files)} files to scan.")
    
    all_entries = {}
    for filepath in files:
        rel_path = os.path.relpath(filepath, ROOT_DIR)
        functions = extract_functions(filepath)
        if functions:
            all_entries[rel_path] = functions

    total_funcs = sum(len(funcs) for funcs in all_entries.values())
    print(f"Detected {total_funcs} functions/methods across active components.")

    commit_id = get_latest_commit_id(ROOT_DIR)

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    sheet_name = "Backend Unit Testing Detail"
    ws = wb.create_sheet(title=sheet_name)
        
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    file_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    font_bold = Font(bold=True)
    
    current_row = 1

    # Add commit ID at the top of the sheet
    ws.cell(row=current_row, column=1, value=f"Commit ID: {commit_id}").font = font_bold
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    current_row += 2
    
    for rel_path, functions in all_entries.items():
        # Add file name row
        ws.cell(row=current_row, column=1, value=f"File: {rel_path}").font = font_bold
        ws.cell(row=current_row, column=1).fill = file_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 1
        
        # Add headers
        headers = ["Sl No.", "Module/Function (Name : Line Number)", "Description", "Actual Input", "Expected Output"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = font_bold
            cell.fill = header_fill
        current_row += 1
        
        # Add data
        sl_no = 1
        for func in functions:
            name_line = f"{func['name']} : {func['line']}"
            ws.cell(row=current_row, column=1, value=sl_no)
            ws.cell(row=current_row, column=2, value=name_line)
            ws.cell(row=current_row, column=3, value=func['description'])
            ws.cell(row=current_row, column=4, value=func['input'])
            ws.cell(row=current_row, column=5, value=func['expected_output'])
            current_row += 1
            sl_no += 1
            
        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40

    wb.save(OUT_FILE)
    print(f"Successfully generated function description sheet at {OUT_FILE}")

if __name__ == '__main__':
    main()
