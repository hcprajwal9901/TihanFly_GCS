import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Target Directories
WORKSPACE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
JS_DIR = os.path.join(WORKSPACE_DIR, 'js')
PLAN_DIR = os.path.join(WORKSPACE_DIR, 'plan-flight-modules')
OUT_FILE = os.path.join(WORKSPACE_DIR, 'test_frontend', 'unit_test', 'all_test_specifications.xlsx')

HEADERS = [
    "File", "Function", "Purpose", "Business Purpose", "Inputs", "Returns", 
    "Dependencies", "Risk Level", "State Changes", "DOM Changes", "WebSocket Messages", 
    "LocalStorage Changes", "Side Effects", "Error Conditions", "Edge Cases", 
    "Branch Conditions", "Happy Path Tests", "Failure Tests", "Boundary Tests", 
    "Required Assertions", "Mock Strategy", "Mutation Checks", "Coverage Priority", 
    "Call Graph", "Analysis Confidence", "Actual Return Types", "Actual Payload Structures", "Called Functions"
]

def extract_function_body(content, start_index):
    """
    Rigorously extracts exact function body matching braces { and }, 
    safely ignoring comments and string literals to prevent brace count mismatches.
    """
    brace_index = content.find('{', start_index)
    if brace_index == -1:
        return ""
    
    count = 1
    i = brace_index + 1
    length = len(content)
    
    in_single_quote = False
    in_double_quote = False
    in_backtick = False
    in_line_comment = False
    in_block_comment = False
    
    while i < length and count > 0:
        char = content[i]
        
        # Track Comment Boundaries
        if in_line_comment:
            if char == '\n':
                in_line_comment = False
        elif in_block_comment:
            if char == '/' and content[i-1] == '*':
                in_block_comment = False
        # Track String Boundaries
        elif in_single_quote:
            if char == "'" and content[i-1] != '\\':
                in_single_quote = False
        elif in_double_quote:
            if char == '"' and content[i-1] != '\\':
                in_double_quote = False
        elif in_backtick:
            if char == '`' and content[i-1] != '\\':
                in_backtick = False
        # Check starting comment/string triggers
        else:
            if char == '/' and i + 1 < length and content[i+1] == '/':
                in_line_comment = True
                i += 1
            elif char == '/' and i + 1 < length and content[i+1] == '*':
                in_block_comment = True
                i += 1
            elif char == "'":
                in_single_quote = True
            elif char == '"':
                in_double_quote = True
            elif char == '`':
                in_backtick = True
            elif char == '{':
                count += 1
            elif char == '}':
                count -= 1
        i += 1
        
    return content[brace_index:i]

def analyze_return_statement(body_code):
    """
    Statically analyzes JS return statements to determine actual return types.
    """
    return_matches = re.findall(r'\breturn\s+(.*?)(?:;|\n|$)', body_code)
    if not return_matches:
        return "void", "None returned (returns undefined implicitly)."
        
    expr = return_matches[0].strip()
    
    if expr in ["true", "false"]:
        return "Boolean", f"Returns Boolean flag `{expr}`."
    elif expr.startswith("[") and expr.endswith("]"):
        return "Array", "Returns inline Array literal."
    elif expr.startswith("{") and expr.endswith("}"):
        return "Object", "Returns inline Object literal."
    elif expr == "null":
        return "null", "Returns null pointer explicitly."
    elif expr.startswith("new Promise") or expr.startswith("Promise."):
        return "Promise", "Returns a Promise representation (Asynchronous queue)."
    elif expr == "this":
        return "Instance (this)", "Returns class instance scope to enable chaining methods."
    elif re.match(r'^[a-zA-Z0-9_\$]+$', expr):
        return f"Object ({expr})", f"Returns local instance variable `{expr}`."
    else:
        return "Dynamic / Inferred", f"Returns evaluated expression: `{expr}`"

def trace_variable_payload(body_code, var_name):
    """
    Traces variables declarations upwards in function body to extract actual object payloads.
    """
    decl_pattern = rf'\b(?:const|let|var)\s+{re.escape(var_name)}\s*=\s*'
    decl_match = re.search(decl_pattern, body_code)
    if not decl_match:
        return ""
    
    start_idx = decl_match.end()
    brace_index = body_code.find('{', start_idx)
    if brace_index == -1:
        # Check if simple literal or call
        end_line_idx = body_code.find(';', start_idx)
        if end_line_idx != -1:
            return body_code[start_idx:end_line_idx].strip()
        return ""
        
    count = 1
    i = brace_index + 1
    length = len(body_code)
    while i < length and count > 0:
        char = body_code[i]
        if char == '{':
            count += 1
        elif char == '}':
            count -= 1
        i += 1
    return body_code[brace_index:i]

def extract_websocket_payloads(body_code):
    """
    Extracts actual object literals passed to WebSocket transmission APIs, including variable tracing.
    """
    ws_regex = r'(?:_wsSend|ws\.send|websocket\.send)\s*\(\s*(\{.*?\})\s*\)'
    payloads = re.findall(ws_regex, body_code, re.DOTALL)
    if not payloads:
        # Check variable send
        var_send = re.findall(r'(?:_wsSend|ws\.send|websocket\.send)\s*\(\s*([a-zA-Z0-9_]+)\s*\)', body_code)
        if var_send:
            var_name = var_send[0]
            traced = trace_variable_payload(body_code, var_name)
            if traced:
                payload_clean = traced.replace('\n', '').replace(' ', '')
                return f"Traced object `{var_name}`: {payload_clean}"
            return f"Variable object: `{var_name}` (Payload shape defined locally)."
        return "No direct WebSocket payload dispatch."
        
    # Clean up inline fields
    payload = payloads[0].replace('\n', '').replace(' ', '')
    return f"Object literal: {payload}"

def parse_javascript_file(filepath):
    """
    Principal-grade static analysis of a JavaScript file using rigorous brace extraction,
    nested call tracing, data-flow payload parsing, and actual return value audits.
    """
    filename = os.path.basename(filepath)
    test_cases = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Regex to locate declarations indexes
        pattern_func = re.finditer(r'function\s+([a-zA-Z0-9_\$]+)\s*\((.*?)\)', content)
        pattern_method = re.finditer(r'([a-zA-Z0-9_\$]+)\s*:\s*(?:function)?\s*\((.*?)\)', content)
        pattern_class_method = re.finditer(r'(?:async\s+)?([a-zA-Z0-9_\$]+)\s*\(([^)]*)\)\s*\{', content)
        
        found_signatures = set()
        raw_declarations = []
        
        for m in pattern_func:
            func_name = m.group(1)
            params = m.group(2)
            start_idx = m.start()
            if func_name not in found_signatures and func_name not in ["if", "for", "switch", "while", "catch"]:
                found_signatures.add(func_name)
                raw_declarations.append((func_name, params, "function", start_idx))
                
        for m in pattern_method:
            func_name = m.group(1)
            params = m.group(2)
            start_idx = m.start()
            if func_name not in found_signatures and func_name not in ["if", "for", "switch", "while", "catch"]:
                found_signatures.add(func_name)
                raw_declarations.append((func_name, params, "class_method", start_idx))
                
        for m in pattern_class_method:
            func_name = m.group(1)
            params = m.group(2)
            start_idx = m.start()
            if func_name not in found_signatures and func_name not in ["if", "for", "switch", "while", "catch", "function", "constructor"]:
                found_signatures.add(func_name)
                raw_declarations.append((func_name, params, "class_method", start_idx))
                
        if "constructor" in content and "constructor" not in found_signatures:
            idx = content.find("constructor")
            raw_declarations.append(("constructor", "", "constructor", idx))
            
        if not raw_declarations:
            raw_declarations.append(("init", "", "event_handler", 0))
            
        # Parse every found function using brace matching and dynamic behavioral mapping
        for func_name, params, func_type, start_idx in raw_declarations:
            # 1. Exact Function Body Extraction
            body_code = extract_function_body(content, start_idx)
            confidence = "High" if len(body_code) > 0 else "Medium"
            if not body_code:
                body_code = content[start_idx:start_idx+1000] # fallback
                confidence = "Low"
                
            # 2. Extract called functions inside the body
            all_calls = re.findall(r'(?:\b([a-zA-Z0-9_\$]+)\s*\()', body_code)
            ignored_words = {
                "if", "for", "switch", "while", "catch", "typeof", "require", 
                "expect", "describe", "it", "beforeEach", "beforeAll", "afterEach", 
                "afterAll", "Promise", "console", "Object", "Array", "String", "Number", 
                "Boolean", "Error", "setTimeout", "setInterval", "parseFloat", "parseInt",
                "Math", "isNaN", "toFixed", "find", "forEach", "map", "filter", "push"
            }
            called_funcs = [c for c in all_calls if c not in ignored_words and c != func_name]
            called_funcs_uniq = sorted(list(set(called_funcs)))
            called_funcs_str = ", ".join(called_funcs_uniq) if called_funcs_uniq else "None"
            
            # 3. Inputs & Returns
            inputs = []
            if params:
                param_names = [p.strip() for p in params.split(',') if p.strip()]
                for p in param_names:
                    p_type = "Number" if p.lower() in ["alt", "altitude", "lat", "lng", "speed", "idx", "id", "val", "x", "y"] else "Object" if p.lower() in ["options", "data", "event", "e"] else "Function" if p.lower() in ["cb", "callback"] else "String"
                    inputs.append(f"{p}: {p_type}")
            inputs_str = ", ".join(inputs) if inputs else "None"
            
            return_type, return_desc = analyze_return_statement(body_code)
            
            # 4. Real Payload Extraction
            ws_payload = extract_websocket_payloads(body_code)
            
            # 5. Extract Detailed DOM Actions
            dom_changes = []
            selections = []
            id_selections = re.findall(r'getElementById\([\'"`](.*?)[\'"`]\)', body_code)
            for id_name in id_selections:
                selections.append((f"#{id_name}", id_name))
            sel_selections = re.findall(r'querySelector\([\'"`](.*?)[\'"`]\)', body_code)
            for sel in sel_selections:
                selections.append((sel, sel))
                
            for sel_label, sel_raw in selections:
                lines_with_sel = [line for line in body_code.split('\n') if sel_raw in line]
                for line in lines_with_sel:
                    if ".click(" in line or ".onclick" in line:
                        dom_changes.append(f"Click element '{sel_label}'")
                    elif ".value" in line or ".val(" in line:
                        dom_changes.append(f"Input value/text into '{sel_label}'")
                    elif "style.display" in line:
                        display_val = re.findall(r'style\.display\s*=\s*[\'"`](.*?)[\'"`]', line)
                        val_str = f" to '{display_val[0]}'" if display_val else ""
                        dom_changes.append(f"Update visibility state of '{sel_label}'{val_str}")
                    elif "classList.add" in line:
                        cls_added = re.findall(r'classList\.add\([\'"`](.*?)[\'"`]\)', line)
                        cls_str = f" '{cls_added[0]}'" if cls_added else ""
                        dom_changes.append(f"Add class{cls_str} to '{sel_label}'")
                    elif "classList.remove" in line:
                        cls_removed = re.findall(r'classList\.remove\([\'"`](.*?)[\'"`]\)', line)
                        cls_str = f" '{cls_removed[0]}'" if cls_removed else ""
                        dom_changes.append(f"Remove class{cls_str} from '{sel_label}'")
                    elif "innerHTML" in line or "textContent" in line:
                        dom_changes.append(f"Update inner HTML/text content of '{sel_label}'")
                        
            if not dom_changes:
                for sel_label, _ in selections:
                    dom_changes.append(f"Query selector '{sel_label}'")
            if "createElement" in body_code:
                created = re.findall(r'createElement\([\'"`](.*?)[\'"`]\)', body_code)
                for c in created:
                    dom_changes.append(f"Create DOM element '<{c}>'")
            if "appendChild" in body_code:
                dom_changes.append("Append child node to container")
                
            dom_changes_str = "\n".join(sorted(list(set(dom_changes)))) if dom_changes else "No direct DOM updates."
            
            # 6. WebSocket Message triggers
            ws_messages = []
            if "ws.send" in body_code or "websocket.send" in body_code or "_wsSend" in body_code:
                ws_messages.append(ws_payload)
            ws_messages_str = "\n".join(ws_messages) if ws_messages else "No network dispatches."
            
            # 7. LocalStorage Caches
            ls_keys = re.findall(r'localStorage\.(?:setItem|getItem|removeItem)\([\'"`](.*?)[\'"`]', body_code)
            ls_keys_str = ", ".join(f"'{k}'" for k in set(ls_keys)) if ls_keys else "No localStorage updates."
            
            # 8. State Mutations
            state_mutations = []
            if "TelemetryStore" in body_code:
                props = re.findall(r'TelemetryStore\.([a-zA-Z0-9_]+)', body_code)
                for p in props:
                    state_mutations.append(f"TelemetryStore.{p}")
            if "this." in body_code:
                props = re.findall(r'this\.([a-zA-Z0-9_]+)\s*=', body_code)
                for p in props:
                    state_mutations.append(f"this.{p}")
            state_str = ", ".join(set(state_mutations)) if state_mutations else "No state updates."
            
            # 9. Errors & branching
            errors = []
            if "throw " in body_code:
                errors.append("Throw raw exception")
            if "console.error" in body_code:
                errors.append("Console diagnostic failure warning")
            if "try" in body_code and "catch" in body_code:
                errors.append("Try-catch fallback guard")
            errors_str = ", ".join(errors) if errors else "Standard exception guard absent."
            
            branches = []
            if "if (" in body_code or "if(" in body_code:
                conds = re.findall(r'if\s*\((.*?)\)', body_code)
                for c in conds[:2]:
                    branches.append(f"Check '{c.strip()}'")
            branches_str = ", ".join(branches) if branches else "Linear execution path."
            
            # 10. Dependency Mapping
            deps = []
            if "L." in body_code or "TMap" in body_code: deps.append("Leaflet API")
            if "waypointManager" in body_code: deps.append("waypointManager")
            if "polygonManager" in body_code: deps.append("polygonManager")
            if "TelemetryStore" in body_code: deps.append("TelemetryStore")
            if "document" in body_code: deps.append("DOM document")
            if "localStorage" in body_code: deps.append("localStorage")
            if "ipcRenderer" in body_code: deps.append("ipcRenderer")
            deps_str = ", ".join(deps) if deps else "Pure algorithmic logic (Independent module)"
            
            # 11. Side Effects
            side_effects = []
            if "addEventListener" in body_code: side_effects.append("Attach event listeners")
            if "setInterval" in body_code or "setTimeout" in body_code: side_effects.append("Timer schedule threads")
            if "ResizeObserver" in body_code: side_effects.append("Layout observation listener")
            se_str = ", ".join(side_effects) if side_effects else "None (Pure computations)."
            
            # 12. Nested Call Graph Tracing
            call_graph = [f"{func_name}()"]
            for call in called_funcs_uniq[:3]: # trace first three invocations
                call_graph.append(f"  -> {call}()")
            if "ws.send" in body_code or "_wsSend" in body_code:
                call_graph.append("  -> WebSocket dispatch")
            call_graph_str = "\n".join(call_graph)
            
            # 13. Safety Risk Tiers & Priority Engine (Code-Aware contextual classification)
            risk = "Low"
            prio = "Low"
            name_lower = func_name.lower()
            file_lower = filename.lower()
            body_lower = body_code.lower()
            
            is_sending_commands = "ws.send" in body_code or "websocket.send" in body_code or "_wssend" in body_lower
            critical_kws = ["takeoff", "land", "rtl", "arm", "disarm", "failsafe", "geofence", "panic", "emergency"]
            high_kws = ["waypoint", "mission", "rally", "telemetry", "compass", "map", "calib", "status", "motor"]
            medium_kws = ["persist", "setting", "config", "i18n", "login", "modal", "panel", "view", "toggle", "window"]
            
            if any(k in name_lower or k in file_lower for k in critical_kws):
                if is_sending_commands or any(k in name_lower for k in ["confirm", "send", "trigger", "write"]):
                    risk = "Critical"
                    prio = "High"
                elif any(k in name_lower for k in ["show", "hide", "modal", "view", "panel", "ui", "display"]):
                    # UI modal wrapper only
                    risk = "Medium"
                    prio = "Medium"
                else:
                    risk = "Critical"
                    prio = "High"
            elif any(k in name_lower or k in file_lower for k in high_kws):
                risk = "High"
                prio = "High"
            elif any(k in name_lower or k in file_lower for k in medium_kws):
                risk = "Medium"
                prio = "Medium"
            elif is_sending_commands:
                risk = "High"
                prio = "High"
                
            # 14. DYNAMIC BEHAVIOR SYNTHESIZER (Based strictly on actual statement tracing!)
            purpose = f"Execute GCS functional logic operations for `{func_name}`."
            business = "Maintains user interface parameters transitions and operations."
            mutation_check = "Invert calculations logic / comparisons."
            happy = f"1. Call {func_name} with target parameters: {inputs_str}.\n2. Verify local operations complete successfully."
            failure = f"1. Pass undefined / invalid parameters to {func_name}.\n2. Verify the system alerts or halts safely."
            boundary = "1. Inject parameters at extreme mathematical bounds.\n2. Verify no buffer overflows or mathematical errors occur."

            # DOM changes behavior synthesis
            if dom_changes:
                target_nodes = ", ".join(set(dom_changes))
                purpose = f"Statically trace and configure view rendering nodes: {target_nodes}."
                business = "Maintains visible layout styling grids and panel configurations for the GCS pilot."
                happy = f"1. Invoke {func_name}({inputs_str if params else ''}).\n2. Verify element selection queries are executed.\n3. Assert target elements are customized successfully."
                failure = "1. Mock target DOM selectors as missing (returns null).\n2. Execute function and assert failure is logged to telemetry console."
                boundary = "1. Trigger rendering on extremely narrow / responsive layout sizes.\n2. Assert GCS layout adjustments prevent visual clipping."
                mutation_check = f"Remove target DOM selections updates to {target_nodes}."

            # WebSocket messages behavior synthesis
            elif ws_messages:
                purpose = f"Formulate serialized payload coordinates and transmit command parameters: {ws_messages_str}."
                business = "Instructs real-time vehicle action triggers and safety alerts on the physical drone."
                happy = f"1. Invoke command trigger `{func_name}` with parameters: {inputs_str}.\n2. Verify WebSocket serialize trigger is executed.\n3. Assert that payload matches structure: {ws_messages_str}."
                failure = f"1. Disconnect network bridge (WebSocket state closed).\n2. Call `{func_name}` and verify SwUtil.toast shows network disconnection alert."
                boundary = f"1. Inject boundary coordinates parameter values.\n2. Verify WebSocket package bounds clamp parameter to GCS standards."
                mutation_check = f"Remove WebSocket transmission invocation inside `{func_name}`."

            # LocalStorage behavior synthesis
            elif ls_keys:
                purpose = f"Persist and store system configuration key-value properties: {ls_keys_str}."
                business = "Enables GCS setting parameters recovery and telemetry states persistence across restarts."
                happy = f"1. Call `{func_name}` to trigger state caching.\n2. Assert that localStorage.setItem is executed with key: {ls_keys_str}."
                failure = "1. Clear storage cache (mock localStorage locked).\n2. Call function and assert recovery procedures default to standard preset configuration."
                boundary = "1. Attempt to cache empty or null configuration values.\n2. Assert standard default configurations write triggers."
                mutation_check = f"Remove localStorage write operation for key {ls_keys_str}."

            # Mathematical / pure calculations behavior synthesis
            elif return_type != "void":
                purpose = f"Calculate functional calculations and evaluate output: `{return_desc}`."
                business = "Provides high precision spatial analytics, mission tracking, and mathematical transformations."
                happy = f"1. Call `{func_name}` with defined inputs: {inputs_str}.\n2. Verify math execution is performed.\n3. Assert that return type is `{return_type}`."
                failure = f"1. Inject invalid type values (undefined / null) into inputs.\n2. Assert that function handles input rejection elegantly or returns default bounds value."
                boundary = "1. Pass parameters at extreme ranges / boundaries limits.\n2. Assert output wraps correctly without buffer errors."
                mutation_check = "Invert logic expressions checks or operators signs."

            # BEHAVIORAL ASSERTIONS SYNTHESIZER (Pulls exact method/function spies!)
            custom_assertions = []
            
            # WebSocket assertions
            if ws_messages:
                custom_assertions.append("expect(wsMock.send).toHaveBeenCalled();")
                if "TAKEOFF" in func_name.upper():
                    custom_assertions.append("expect(wsMock.send).toHaveBeenCalledWith(expect.stringContaining('\"type\":\"TAKEOFF\"'));")
                elif "LAND" in func_name.upper():
                    custom_assertions.append("expect(wsMock.send).toHaveBeenCalledWith(expect.stringContaining('\"type\":\"LAND\"'));")
                elif "RTL" in func_name.upper():
                    custom_assertions.append("expect(wsMock.send).toHaveBeenCalledWith(expect.stringContaining('\"type\":\"RTL\"'));")
            
            # LocalStorage assertions
            if ls_keys:
                custom_assertions.append(f"expect(localStorage.setItem).toHaveBeenCalledWith({ls_keys_str}, expect.any(String));")
                
            # DOM mutations assertions
            if dom_changes:
                if "classList" in dom_changes_str:
                    custom_assertions.append("expect(element.className).toContain('active');")
                if "style" in dom_changes_str:
                    custom_assertions.append("expect(element.style.display).toBeDefined();")
            
            # Called functions and browser APIs assertions
            for call in called_funcs_uniq:
                # Ignore noise functions (utility, logger, and DOM helpers)
                if call.lower() in [
                    "querySelector", "getelementbyid", "classlist", "add", "remove", "toggle", 
                    "_wssend", "ws.send", "websocket.send", "setitem", "getitem", "removeitem",
                    "log", "logger", "console.log", "console.error", "console.warn",
                    "msgconsole.log", "msgconsole.error", "msgconsole.success", "msgconsole.warn", "msgconsole.info",
                    "buildpanelhtml", "error", "success", "info"
                ]:
                    continue
                
                if call in ["createObjectURL", "revokeObjectURL"]:
                    custom_assertions.append("expect(URL.createObjectURL).toHaveBeenCalled();")
                elif call in ["createElement"]:
                    custom_assertions.append("expect(document.createElement).toHaveBeenCalledWith('a');")
                elif call in ["click"] and "createObjectURL" in called_funcs_uniq:
                    custom_assertions.append("expect(anchor.click).toHaveBeenCalled();")
                elif call in ["toast", "SwUtil.toast"]:
                    custom_assertions.append("expect(SwUtil.toast).toHaveBeenCalled();")
                elif call in ["takeoff", "land", "rtl"]:
                    custom_assertions.append(f"expect(MsgConsole.{call}).toHaveBeenCalled();")
                else:
                    custom_assertions.append(f"expect({call}).toHaveBeenCalled();")
                    
            if not custom_assertions:
                if return_type != "void":
                    custom_assertions.append(f"const res = {func_name}(...);")
                    custom_assertions.append("expect(res).toBeDefined();")
                    if "Boolean" in return_type:
                        custom_assertions.append("expect(typeof res).toBe('boolean');")
                    elif "Array" in return_type:
                        custom_assertions.append("expect(Array.isArray(res)).toBe(true);")
                else:
                    custom_assertions.append(f"expect({func_name}).toBeDefined();")
                    
            assertions = "\n".join(sorted(list(set(custom_assertions))))

            mock_strategy = "Live execution."
            mocks = []
            if "Leaflet" in deps_str or "Map" in deps_str: mocks.append("Mock L Map layers")
            if ws_messages: mocks.append("Mock WebSocket connections")
            if ls_keys: mocks.append("Mock localStorage cache items")
            if "ipcRenderer" in deps_str: mocks.append("Mock Electron IPC channels")
            if mocks:
                mock_strategy = f"Mock external boundaries: {', '.join(mocks)}."
                
            test_cases.append({
                "file": filename,
                "function": func_name,
                "purpose": purpose,
                "business_purpose": business,
                "inputs": inputs_str,
                "returns": return_desc,
                "dependencies": deps_str,
                "risk_level": risk,
                "state_changes": state_str,
                "dom_changes": dom_changes_str,
                "websocket_messages": ws_messages_str,
                "localstorage_changes": ls_keys_str,
                "side_effects": se_str,
                "error_conditions": errors_str,
                "edge_cases": "Missing/empty coordinates inputs or WebSocket network dropouts.",
                "branch_conditions": branches_str,
                "happy_path": happy,
                "failure_path": failure,
                "boundary_path": boundary,
                "assertions": assertions,
                "mock_strategy": mock_strategy,
                "mutation_checks": mutation_check,
                "coverage_priority": prio,
                "call_graph": call_graph_str,
                "analysis_confidence": confidence,
                "actual_return_types": return_type,
                "actual_payload_structures": ws_messages_str,
                "called_functions": called_funcs_str
            })
            
    except Exception as e:
        print(f"Error statically scanning file {filename}: {e}")
        
    return test_cases

def main():
    print("==================================================")
    print("STARTING PRINCIPAL-GRADE TEST SPECIFICATION ENGINE")
    print("==================================================")
    
    all_specifications = []
    
    # 1. Scan plan-flight-modules
    if os.path.exists(PLAN_DIR):
        print(f"Scanning directory: {PLAN_DIR}")
        for file in os.listdir(PLAN_DIR):
            if file.endswith('.js'):
                fullpath = os.path.join(PLAN_DIR, file)
                all_specifications.extend(parse_javascript_file(fullpath))
                
    # 2. Scan js directory
    if os.path.exists(JS_DIR):
        print(f"Scanning directory: {JS_DIR}")
        for file in os.listdir(JS_DIR):
            if file.endswith('.js'):
                fullpath = os.path.join(JS_DIR, file)
                all_specifications.extend(parse_javascript_file(fullpath))
                
    print(f"Successfully compiled principal specifications for {len(all_specifications)} functions.")
    
    # Build openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Specifications"
    
    # Fonts and fills
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_data = Font(name="Segoe UI", size=9)
    
    fill_title = PatternFill(start_color="1E1E3F", end_color="1E1E3F", fill_type="solid") # Dark premium blue
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # silver
    fill_stripe = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid") # zebra stripe
    
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Merged title row
    ws.cell(row=1, column=1, value="TiHANFly GCS Frontend Unit Testing Specification Dataset").font = font_title
    ws.cell(row=1, column=1).fill = fill_title
    ws.row_dimensions[1].height = 40
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=28)
    
    ws.row_dimensions[2].height = 10 # empty divider
    
    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    ws.row_dimensions[3].height = 30
    
    # Add Rows
    current_row = 4
    for spec in all_specifications:
        # Write exact 28 columns mapping
        ws.cell(row=current_row, column=1, value=spec["file"])
        ws.cell(row=current_row, column=2, value=spec["function"])
        ws.cell(row=current_row, column=3, value=spec["purpose"])
        ws.cell(row=current_row, column=4, value=spec["business_purpose"])
        ws.cell(row=current_row, column=5, value=spec["inputs"])
        ws.cell(row=current_row, column=6, value=spec["returns"])
        ws.cell(row=current_row, column=7, value=spec["dependencies"])
        ws.cell(row=current_row, column=8, value=spec["risk_level"])
        ws.cell(row=current_row, column=9, value=spec["state_changes"])
        ws.cell(row=current_row, column=10, value=spec["dom_changes"])
        ws.cell(row=current_row, column=11, value=spec["websocket_messages"])
        ws.cell(row=current_row, column=12, value=spec["localstorage_changes"])
        ws.cell(row=current_row, column=13, value=spec["side_effects"])
        ws.cell(row=current_row, column=14, value=spec["error_conditions"])
        ws.cell(row=current_row, column=15, value=spec["edge_cases"])
        ws.cell(row=current_row, column=16, value=spec["branch_conditions"])
        ws.cell(row=current_row, column=17, value=spec["happy_path"])
        ws.cell(row=current_row, column=18, value=spec["failure_path"])
        ws.cell(row=current_row, column=19, value=spec["boundary_path"])
        ws.cell(row=current_row, column=20, value=spec["assertions"])
        ws.cell(row=current_row, column=21, value=spec["mock_strategy"])
        ws.cell(row=current_row, column=22, value=spec["mutation_checks"])
        ws.cell(row=current_row, column=23, value=spec["coverage_priority"])
        ws.cell(row=current_row, column=24, value=spec["call_graph"])
        ws.cell(row=current_row, column=25, value=spec["analysis_confidence"])
        ws.cell(row=current_row, column=26, value=spec["actual_return_types"])
        ws.cell(row=current_row, column=27, value=spec["actual_payload_structures"])
        ws.cell(row=current_row, column=28, value=spec["called_functions"])
        
        # Apply formatting and zebra striping
        for col_idx in range(1, 29):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            # Alignments format
            if col_idx in [1, 2, 8, 23, 25, 26]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
            if current_row % 2 == 0:
                cell.fill = fill_stripe
                
        ws.row_dimensions[current_row].height = 45
        current_row += 1
        
    # Auto adjust column widths elegantly
    widths = {
        'A': 16, 'B': 22, 'C': 30, 'D': 35, 'E': 15, 'F': 22, 'G': 18, 'H': 12,
        'I': 22, 'J': 25, 'K': 25, 'L': 22, 'M': 20, 'N': 20, 'O': 20, 'P': 22,
        'Q': 35, 'R': 35, 'S': 35, 'T': 30, 'U': 30, 'V': 30, 'W': 12, 'X': 25,
        'Y': 14, 'Z': 16, 'AA': 25, 'AB': 25
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
        
    # Save Workbook with lock checks
    try:
        wb.save(OUT_FILE)
        print("==================================================")
        print(f"SUCCESS: Compiled Principal 28-Column Specifications to Excel:")
        print(f"--> {OUT_FILE}")
        print("==================================================")
    except PermissionError:
        fallback_file = OUT_FILE.replace('.xlsx', '_v2.xlsx')
        wb.save(fallback_file)
        print("==================================================")
        print(f"WARNING: The target file {OUT_FILE} is locked.")
        print(f"Fallback successfully saved compiled specifications to:")
        print(f"--> {fallback_file}")
        print("==================================================")

if __name__ == '__main__':
    main()
