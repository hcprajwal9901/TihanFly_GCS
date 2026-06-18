import os
import json
import re
import platform
import subprocess
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
JSON_PATH = os.path.join(os.path.dirname(__file__), 'test-results.json')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'reports', 'integration_test_report.xlsx')
HISTORY_PATH = os.path.join(os.path.dirname(__file__), 'reports', 'history.json')

def sanitize_string(val):
    if not isinstance(val, str):
        return val
    # Remove ANSI escape sequences
    val = re.sub(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])', '', val)
    # Remove XML-illegal characters (except valid whitespaces)
    val = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', val)
    return val

def classify_test(tc):
    title = (tc.get('suite', '') + ' ' + tc.get('name', '')).lower()
    
    if 'accessibility' in title or 'wcag' in title or 'axe' in title:
        return 'Accessibility', 'Minor', 'No'
    elif 'visual' in title or 'screenshot' in title or 'regression' in title or 'toHaveScreenshot' in title:
        return 'Visual Regression', 'Minor', 'No'
    elif 'performance' in title or 'benchmark' in title or 'timing' in title or 'switch' in title or 'duration' in title:
        return 'Performance', 'Major', 'Yes'
    elif 'contract' in title or 'preload' in title or 'whitelist' in title or 'apis' in title or 'require' in title:
        return 'IPC Contract', 'Critical', 'Yes'
    elif 'ack timeout' in title or 'reliability' in title or 'double-click' in title:
        return 'Command Reliability', 'Critical', 'Yes'
    elif 'reconnect' in title or 'recovery' in title or 'out-of-order' in title or 'duplicate' in title:
        return 'Telemetry Recovery', 'Critical', 'Yes'
    elif 'geofence' in title or 'fence' in title:
        return 'Geofence', 'Critical', 'Yes'
    elif 'firmware' in title or 'flash' in title:
        return 'Firmware Recovery', 'Critical', 'Yes'
    elif 'telemetry' in title or 'scenario' in title or 'loss' in title or 'gps' in title or 'battery' in title or 'packet' in title or 'network' in title:
        return 'Telemetry', 'Critical', 'Yes'
    elif 'workflow' in title or 'arm' in title or 'state-machine' in title or 'rtl' in title or 'takeoff' in title or 'disarm' in title:
        return 'Flight Workflow', 'Critical', 'Yes'
    elif 'sitl' in title or 'simulation' in title:
        return 'SITL Validation', 'Critical', 'Yes'
    else:
        return 'Navigation/UI', 'Major', 'Yes'

def main():
    print("[Report] Parsing Playwright JSON test results...")
    
    # Load KPIs if present
    kpi_data = None
    kpi_path = os.path.join(os.path.dirname(__file__), 'reports', 'kpis.json')
    if os.path.exists(kpi_path):
        try:
            with open(kpi_path, 'r', encoding='utf-8') as kf:
                kpi_data = json.load(kf)
            print("[Report] Loaded KPIs successfully.")
        except Exception as e:
            print(f"[Warning] Failed to read kpis.json: {e}")

    if not os.path.exists(JSON_PATH):
        print(f"Error: JSON results file not found at {JSON_PATH}. Run tests first.")
        return

    try:
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error reading JSON file: {e}")
        return

    # Extract statistics
    stats = data.get('stats', {})
    total_duration_ms = stats.get('duration', 0)
    total_duration_sec = total_duration_ms / 1000
    minutes = int(total_duration_sec // 60)
    seconds = int(total_duration_sec % 60)
    duration_str = f"{minutes:02d}:{seconds:02d}"

    # Extract test cases
    test_cases = []
    
    def parse_suite(suite, file_name):
        suite_title = suite.get('title', '')
        for spec in suite.get('specs', []):
            spec_title = spec.get('title', '')
            for test_item in spec.get('tests', []):
                results = test_item.get('results', [])
                if not results:
                    continue
                
                result = results[-1]
                raw_status = result.get('status', '')
                
                if raw_status == 'passed' or test_item.get('status') == 'expected':
                    status = 'PASS'
                elif raw_status == 'skipped' or test_item.get('status') == 'skipped':
                    status = 'SKIPPED'
                else:
                    status = 'FAIL'

                duration = result.get('duration', 0)
                error_msg = ""
                error = result.get('error')
                if error:
                    error_msg = sanitize_string(error.get('message', ''))
                
                # Extract stdout and stderr
                log_lines = []
                for out_item in result.get('stdout', []):
                    text = out_item.get('text', '')
                    if text:
                        log_lines.append(text)
                for err_item in result.get('stderr', []):
                    text = err_item.get('text', '')
                    if text:
                        log_lines.append(text)

                attachments = []
                evidence_dir = os.path.join(os.path.dirname(EXCEL_PATH), 'evidence')
                os.makedirs(evidence_dir, exist_ok=True)
                
                safe_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', spec_title)[:60]
                test_num = len(test_cases) + 1

                # Copy existing attachments (screenshots, context files)
                for att in result.get('attachments', []):
                    att_path = att.get('path', '')
                    if att_path and os.path.exists(att_path):
                        dest_filename = f"{test_num}_{safe_name}_{os.path.basename(att_path)}"
                        dest_path = os.path.join(evidence_dir, dest_filename)
                        
                        try:
                            import shutil
                            shutil.copy2(att_path, dest_path)
                            rel_path = f"evidence/{dest_filename}"
                            
                            att_name = att.get('name', '').lower()
                            if att_path.lower().endswith('.png') or 'screenshot' in att_name:
                                att_type = 'screenshot'
                            elif att_path.lower().endswith('.md') or 'context' in att_name:
                                att_type = 'context'
                            else:
                                att_type = 'context'

                            attachments.append({
                                'name': att.get('name', 'Evidence'),
                                'path': rel_path,
                                'type': att_type
                            })
                        except Exception as e:
                            print(f"[Warning] Failed to copy evidence file {att_path}: {e}")

                # Save execution logs to evidence folder
                log_filename = f"{test_num}_{safe_name}_logs.txt"
                log_path = os.path.join(evidence_dir, log_filename)
                try:
                    with open(log_path, 'w', encoding='utf-8') as log_f:
                        if log_lines:
                            log_f.writelines(log_lines)
                        else:
                            log_f.write("No console logs captured for this test case.")
                    attachments.append({
                        'name': 'Test Logs',
                        'path': f"evidence/{log_filename}",
                        'type': 'logs'
                    })
                except Exception as e:
                    print(f"[Warning] Failed to write test log {log_path}: {e}")

                test_cases.append({
                    'file': file_name,
                    'suite': suite_title,
                    'name': spec_title,
                    'status': status,
                    'duration': duration,
                    'error': error_msg,
                    'attachments': attachments
                })

        for sub_suite in suite.get('suites', []):
            parse_suite(sub_suite, file_name or suite.get('file', ''))

    for root_suite in data.get('suites', []):
        file_name = root_suite.get('file', '')
        parse_suite(root_suite, file_name)

    total_tests = len(test_cases)
    passed_tests = sum(1 for tc in test_cases if tc['status'] == 'PASS')
    failed_tests = sum(1 for tc in test_cases if tc['status'] == 'FAIL')
    skipped_tests = sum(1 for tc in test_cases if tc['status'] == 'SKIPPED')
    success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 100.0

    # Determine Release Gate Status
    is_blocked = False
    failure_counts_by_cat = {
        'Flight Workflow': 0,
        'IPC Contract': 0,
        'Telemetry': 0,
        'Performance': 0,
        'Navigation/UI': 0,
        'Visual Regression': 0,
        'Accessibility': 0,
        'Command Reliability': 0,
        'Telemetry Recovery': 0,
        'Geofence': 0,
        'Firmware Recovery': 0,
        'SITL Validation': 0
    }
    
    for tc in test_cases:
        category, severity, blocks = classify_test(tc)
        if tc['status'] == 'FAIL':
            failure_counts_by_cat[category] += 1
            if blocks == 'Yes':
                is_blocked = True

    # Calculate Flaky Rate
    unique_tests = {}
    for tc in test_cases:
        key = (tc['file'], tc['suite'], tc['name'])
        if key not in unique_tests:
            unique_tests[key] = []
        unique_tests[key].append(tc['status'])

    flaky_count = 0
    for key, statuses in unique_tests.items():
        has_pass = 'PASS' in statuses
        has_fail = 'FAIL' in statuses
        if has_pass and has_fail:
            flaky_count += 1

    total_unique = len(unique_tests)
    flaky_rate = (flaky_count / total_unique * 100.0) if total_unique > 0 else 0.0

    # Check for KPI status
    kpi_failures = False
    if kpi_data:
        metrics_kpi = kpi_data.get("metrics", {})
        for metric_name, m_info in metrics_kpi.items():
            if m_info.get("pass") is False:
                kpi_failures = True
        
        # Also check reliability section
        reliability = kpi_data.get("reliability", {})
        for comp, rel_status in reliability.items():
            if rel_status == "FAIL":
                kpi_failures = True

    # Check for MAVLink Gate Status
    mavlink_gate = "APPROVED" # default
    mavlink_compliance = None
    mavlink_loss = None
    mavlink_crc = None
    
    if kpi_data and "mavlink" in kpi_data:
        m_stats = kpi_data["mavlink"]
        mavlink_compliance = m_stats.get("compliance_score", 0.0)
        mavlink_loss = m_stats.get("packet_loss_pct", 0.0)
        mavlink_crc = m_stats.get("crc_failures", 0)
        
        # Gates enforcement
        if mavlink_compliance < 90.0 or mavlink_crc > 0 or mavlink_loss > 3.0:
            mavlink_gate = "BLOCKED"
        elif mavlink_compliance < 95.0 or mavlink_loss >= 1.0:
            mavlink_gate = "APPROVED WITH PROTOCOL DEBT"

    # Determine Release Gate Status
    if is_blocked or mavlink_gate == "BLOCKED":
        release_status = "BLOCKED"
    elif mavlink_gate == "APPROVED WITH PROTOCOL DEBT":
        release_status = "APPROVED WITH PROTOCOL DEBT"
    elif kpi_failures:
        release_status = "APPROVED WITH PERFORMANCE DEBT"
    else:
        release_status = "APPROVED"

    # Environment Metadata Gathering
    os_info = f"{platform.system()} {platform.release()}"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        git_commit = subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD'], stderr=subprocess.DEVNULL).decode('utf-8').strip()
    except Exception:
        git_commit = 'N/A'

    try:
        git_branch = subprocess.check_output(['git', 'rev-parse', '--abbrev-ref', 'HEAD'], stderr=subprocess.DEVNULL).decode('utf-8').strip()
    except Exception:
        git_branch = 'N/A'

    electron_ver = 'N/A'
    playwright_ver = 'N/A'
    try:
        with open(os.path.join(os.path.dirname(__file__), '../package.json'), 'r') as pkg_f:
            pkg_data = json.load(pkg_f)
            electron_ver = pkg_data.get('devDependencies', {}).get('electron', 'N/A')
            playwright_ver = pkg_data.get('devDependencies', {}).get('@playwright/test', 'N/A')
    except Exception:
        pass

    # Save to history.json (Persistent trend logs)
    history_data = []
    if os.path.exists(HISTORY_PATH):
        try:
            with open(HISTORY_PATH, 'r', encoding='utf-8') as hist_f:
                history_data = json.load(hist_f)
        except Exception:
            pass

    next_build_num = len(history_data) + 1
    # Check if we are running in GHA
    gha_build = os.environ.get('GITHUB_RUN_NUMBER')
    build_label = f"#{gha_build}" if gha_build else f"Local-{next_build_num}"

    history_data.append({
        "build": build_label,
        "timestamp": timestamp,
        "total": total_tests,
        "passed": passed_tests,
        "failed": failed_tests,
        "skipped": skipped_tests,
        "pass_rate": round(success_rate, 2),
        "flaky_rate": round(flaky_rate, 2),
        "status": release_status,
        "git_commit": git_commit,
        "branch": git_branch,
        "duration_ms": total_duration_ms
    })

    os.makedirs(os.path.dirname(HISTORY_PATH), exist_ok=True)
    with open(HISTORY_PATH, 'w', encoding='utf-8') as hist_w:
        json.dump(history_data, hist_w, indent=2)

    # Excel Style Tokens
    font_family = 'Segoe UI'
    title_font = Font(name=font_family, size=16, bold=True, color='1A2332')
    section_font = Font(name=font_family, size=12, bold=True, color='1A2332')
    header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    link_font = Font(name=font_family, size=11, color='0056B3', underline='single')
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') 
    accent_fill = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid') 
    pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid') 
    fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid') 
    skip_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid') 

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    # Title
    ws1.cell(row=2, column=2, value="TiHAN GCS - Integration Testing Summary").font = title_font

    # Release Gate Status Banner
    status_cell = ws1.cell(row=4, column=2, value=f"RELEASE STATUS: {release_status}")
    status_cell.font = Font(name=font_family, size=14, bold=True, color='FFFFFF')
    status_cell.alignment = Alignment(horizontal='center', vertical='center')
    if release_status == 'APPROVED':
        status_fill_color = '28A745'
    elif release_status == 'APPROVED WITH PERFORMANCE DEBT':
        status_fill_color = 'FD7E14'
    elif release_status == 'APPROVED WITH PROTOCOL DEBT':
        status_fill_color = '6F42C1' # Purple/Violet
    else:
        status_fill_color = 'DC3545'
    status_cell.fill = PatternFill(start_color=status_fill_color, end_color=status_fill_color, fill_type='solid')
    ws1.merge_cells("B4:E4")
    ws1.row_dimensions[4].height = 30

    # Environment Metadata Box
    ws1.cell(row=6, column=2, value="Execution Environment Details").font = section_font
    metadata_fields = [
        ("Operating System", os_info),
        ("Git Commit Reference", git_commit),
        ("Electron Version", electron_ver),
        ("Playwright Version", playwright_ver),
        ("Execution Timestamp", timestamp),
        ("Terminal Execution Log", "View Log")
    ]
    for idx, (lbl, val) in enumerate(metadata_fields, 7):
        c1 = ws1.cell(row=idx, column=2, value=lbl)
        c1.font = bold_font
        c1.fill = accent_fill
        c1.border = thin_border
        
        c2 = ws1.cell(row=idx, column=3, value=val)
        c2.border = thin_border
        if lbl == "Terminal Execution Log":
            c2.hyperlink = "evidence/playwright_execution.log"
            c2.font = link_font
        else:
            c2.font = regular_font
        ws1.row_dimensions[idx].height = 20

    # Summary Metrics Table
    ws1.cell(row=6, column=5, value="Run Statistics").font = section_font
    metrics = [
        ("Total Tests Run", total_tests),
        ("Passed Tests", passed_tests),
        ("Failed Tests", failed_tests),
        ("Skipped Tests", skipped_tests),
        ("Success Rate", f"{success_rate:.2f}%"),
        ("Flaky Rate", f"{flaky_rate:.2f}%"),
        ("Execution Duration", duration_str)
    ]
    for idx, (lbl, val) in enumerate(metrics, 7):
        c1 = ws1.cell(row=idx, column=5, value=lbl)
        c1.font = bold_font
        c1.fill = accent_fill
        c1.border = thin_border

        c2 = ws1.cell(row=idx, column=6, value=val)
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='right')
        
        if lbl == "Success Rate":
            if success_rate >= 90.0:
                c2.fill = pass_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='155724')
            else:
                c2.fill = fail_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        elif lbl == "Flaky Rate":
            if flaky_rate == 0.0:
                c2.fill = pass_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='155724')
            else:
                c2.fill = fail_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        elif lbl == "Failed Tests" and failed_tests > 0:
            c2.fill = fail_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        ws1.row_dimensions[idx].height = 20

    # Failure Classification Table
    ws1.cell(row=14, column=2, value="Failure Classification Summary").font = section_font
    headers = ["Category", "Failure Count"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws1.cell(row=15, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    ws1.row_dimensions[15].height = 20

    row_pos = 16
    for cat, count in failure_counts_by_cat.items():
        c1 = ws1.cell(row=row_pos, column=2, value=cat)
        c1.font = regular_font
        c1.border = thin_border

        c2 = ws1.cell(row=row_pos, column=3, value=count)
        c2.font = bold_font if count > 0 else regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='center')
        if count > 0:
            c2.fill = fail_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        ws1.row_dimensions[row_pos].height = 20
        row_pos += 1

    # ─── KPI Section (E14:H21) ───
    ws1.cell(row=14, column=5, value="Key Performance Indicators (KPIs)").font = section_font
    kpi_headers = ["KPI Metric", "Target", "Measured Value", "Status"]
    for col_offset, h in enumerate(kpi_headers, 5):
        cell = ws1.cell(row=15, column=col_offset, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    ws1.row_dimensions[15].height = 20

    # Parse KPI data
    rec_time_str = "—"
    rec_status = "SKIP"
    ack_time_str = "—"
    ack_status = "SKIP"
    gf_time_str = "—"
    gf_status = "SKIP"
    
    rec_success_rate = "—"
    ack_success_rate = "—"
    telem_recovery_rate = "—"

    if kpi_data:
        m = kpi_data.get("metrics", {})
        
        # Reconnect Recovery Time
        r_m = m.get("reconnect_time_ms", {})
        if r_m.get("measured", 0) > 0:
            rec_time_str = f"{(r_m['measured']/1000.0):.2f} s"
            rec_status = "PASS" if r_m.get("pass", False) else "FAIL"
            
        # ACK Timeout Detection Time
        a_m = m.get("ack_timeout_detection_ms", {})
        if a_m.get("measured", 0) > 0:
            ack_time_str = f"{(a_m['measured']/1000.0):.2f} s"
            ack_status = "PASS" if a_m.get("pass", False) else "FAIL"
            
        # Geofence Warning Latency
        g_m = m.get("geofence_warning_latency_ms", {})
        if g_m.get("measured", 0) > 0:
            gf_time_str = f"{(g_m['measured']/1000.0):.2f} s"
            gf_status = "PASS" if g_m.get("pass", False) else "FAIL"

        # Success Rates
        rec_success_rate = "100.0%" if kpi_data.get("reliability", {}).get("reconnect_recovery") == "PASS" else "0.0%"
        ack_success_rate = "100.0%" if kpi_data.get("reliability", {}).get("ack_timeout_recovery") == "PASS" else "0.0%"
        telem_recovery_rate = "100.0%" if kpi_data.get("reliability", {}).get("out_of_order_handling") == "PASS" else "0.0%"

    kpi_rows = [
        ("Reconnect Recovery Time", "< 2.0 s", rec_time_str, rec_status),
        ("ACK Timeout Detection Time", "< 5.0 s", ack_time_str, ack_status),
        ("Geofence Warning Latency", "< 1.0 s", gf_time_str, gf_status),
        ("Reconnect Success Rate", "100.0%", rec_success_rate, "PASS" if rec_success_rate == "100.0%" else "FAIL"),
        ("ACK Timeout Detection Rate", "100.0%", ack_success_rate, "PASS" if ack_success_rate == "100.0%" else "FAIL"),
        ("Telemetry Recovery Success Rate", "100.0%", telem_recovery_rate, "PASS" if telem_recovery_rate == "100.0%" else "FAIL"),
    ]

    for idx, (metric_name, target_val, measured_val, status) in enumerate(kpi_rows, 16):
        c1 = ws1.cell(row=idx, column=5, value=metric_name)
        c2 = ws1.cell(row=idx, column=6, value=target_val)
        c3 = ws1.cell(row=idx, column=7, value=measured_val)
        c4 = ws1.cell(row=idx, column=8, value=status)

        c1.font = regular_font
        c2.font = regular_font
        c3.font = regular_font
        c4.font = bold_font

        c2.alignment = Alignment(horizontal='center')
        c3.alignment = Alignment(horizontal='right')
        c4.alignment = Alignment(horizontal='center')

        for cell in [c1, c2, c3, c4]:
            cell.border = thin_border

        if status == "PASS":
            c4.fill = pass_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='155724')
        elif status == "FAIL":
            c4.fill = fail_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='721C24')
        else:
            c4.fill = skip_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='856404')
        ws1.row_dimensions[idx].height = 20

    # ─── Operational Reliability Summary Section (E23:F29) ───
    ws1.cell(row=23, column=5, value="Operational Reliability Summary").font = section_font
    rel_headers = ["Reliability Component", "Result"]
    for col_offset, h in enumerate(rel_headers, 5):
        cell = ws1.cell(row=24, column=col_offset, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    ws1.row_dimensions[24].height = 20

    ack_val = kpi_data.get("reliability", {}).get("ack_timeout_recovery", "—") if kpi_data else "—"
    rec_val = kpi_data.get("reliability", {}).get("reconnect_recovery", "—") if kpi_data else "—"
    dup_val = kpi_data.get("reliability", {}).get("duplicate_packet_handling", "—") if kpi_data else "—"
    ooo_val = kpi_data.get("reliability", {}).get("out_of_order_handling", "—") if kpi_data else "—"
    gps_val = kpi_data.get("reliability", {}).get("gps_recovery_handling", "—") if kpi_data else "—"

    reliability_rows = [
        ("ACK Timeout Recovery", ack_val),
        ("Reconnect Recovery", rec_val),
        ("Duplicate Packet Handling", dup_val),
        ("Out-of-Order Handling", ooo_val),
        ("GPS Recovery Handling", gps_val),
    ]

    for idx, (comp_name, result) in enumerate(reliability_rows, 25):
        c1 = ws1.cell(row=idx, column=5, value=comp_name)
        c2 = ws1.cell(row=idx, column=6, value=result)

        c1.font = regular_font
        c2.font = bold_font
        c2.alignment = Alignment(horizontal='center')

        c1.border = thin_border
        c2.border = thin_border

        if result == "PASS":
            c2.fill = pass_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='155724')
        elif result == "FAIL":
            c2.fill = fail_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        else:
            c2.fill = skip_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='856404')
        ws1.row_dimensions[idx].height = 20

    # ════════════════════════════════════════════════
    # SHEET 2: DETAILED RESULTS
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Detailed Results")
    ws2.views.sheetView[0].showGridLines = True

    headers = ["Test File", "Suite/Describe", "Test Case", "Status", "Duration (ms)", "Category", "Severity", "Screenshot / Snapshot", "Test Execution Logs", "Error Context"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if header in ["Status", "Duration (ms)", "Severity", "Screenshot / Snapshot", "Test Execution Logs", "Error Context"] else 'left')
        cell.border = thin_border
    ws2.row_dimensions[2].height = 24

    for row_idx, tc in enumerate(test_cases, 3):
        clean_file = os.path.basename(tc['file'])
        cat, sev, blocks = classify_test(tc)
        
        c1 = ws2.cell(row=row_idx, column=1, value=clean_file)
        c2 = ws2.cell(row=row_idx, column=2, value=tc['suite'])
        c3 = ws2.cell(row=row_idx, column=3, value=tc['name'])
        
        c4 = ws2.cell(row=row_idx, column=4, value=tc['status'])
        c4.alignment = Alignment(horizontal='center')
        if tc['status'] == 'PASS':
            c4.fill = pass_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='155724')
        elif tc['status'] == 'FAIL':
            c4.fill = fail_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='721C24')
        else:
            c4.fill = skip_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='856404')

        c5 = ws2.cell(row=row_idx, column=5, value=tc['duration'])
        c5.alignment = Alignment(horizontal='right')

        c6 = ws2.cell(row=row_idx, column=6, value=cat)
        
        c7 = ws2.cell(row=row_idx, column=7, value=sev)
        c7.alignment = Alignment(horizontal='center')
        if tc['status'] == 'FAIL':
            if sev == 'Critical':
                c7.font = Font(name=font_family, size=11, bold=True, color='721C24')
                c7.fill = fail_fill

        # Extract specific attachments by category
        screenshot_att = None
        logs_att = None
        context_att = None
        for att in tc.get('attachments', []):
            t = att.get('type')
            if t == 'screenshot':
                screenshot_att = att
            elif t == 'logs':
                logs_att = att
            elif t == 'context':
                context_att = att

        # Column 8: Screenshot / Snapshot
        c8 = ws2.cell(row=row_idx, column=8)
        c8.alignment = Alignment(horizontal='center')
        if screenshot_att:
            c8.value = "View Screenshot"
            c8.hyperlink = screenshot_att['path']
            c8.font = link_font
        else:
            c8.value = "—"
            c8.font = regular_font

        # Column 9: Test Execution Logs
        c9 = ws2.cell(row=row_idx, column=9)
        c9.alignment = Alignment(horizontal='center')
        if logs_att:
            c9.value = "View Logs"
            c9.hyperlink = logs_att['path']
            c9.font = link_font
        else:
            c9.value = "—"
            c9.font = regular_font

        # Column 10: Error Context
        c10 = ws2.cell(row=row_idx, column=10)
        c10.alignment = Alignment(horizontal='center')
        if context_att:
            c10.value = "View Context"
            c10.hyperlink = context_att['path']
            c10.font = link_font
        else:
            c10.value = "—"
            c10.font = regular_font

        for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10]:
            cell.border = thin_border
            if cell not in [c4, c7]:
                cell.font = regular_font
        
        ws2.row_dimensions[row_idx].height = 20

    # ════════════════════════════════════════════════
    # SHEET 3: FAILURE ANALYSIS
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet(title="Failure Analysis")
    ws3.views.sheetView[0].showGridLines = True

    ws3.cell(row=2, column=1, value="Failed Test Case Diagnostics").font = title_font
    
    headers = ["Test Case", "Category", "Severity", "Diagnostic Error Message", "Screenshot", "Test Logs", "Error Context"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if header in ["Severity", "Screenshot", "Test Logs", "Error Context"] else 'left')
        cell.border = thin_border
    ws3.row_dimensions[4].height = 24

    failed_tc_list = [tc for tc in test_cases if tc['status'] == 'FAIL']
    
    if not failed_tc_list:
        ws3.cell(row=5, column=1, value="All tests passed successfully! No failures detected.").font = regular_font
        ws3.merge_cells("A5:G5")
        ws3.row_dimensions[5].height = 22
    else:
        for idx, tc in enumerate(failed_tc_list, 5):
            cat, sev, blocks = classify_test(tc)
            c1 = ws3.cell(row=idx, column=1, value=f"{tc['suite']} > {tc['name']}")
            c2 = ws3.cell(row=idx, column=2, value=cat)
            
            c3 = ws3.cell(row=idx, column=3, value=sev)
            c3.alignment = Alignment(horizontal='center')
            if sev == 'Critical':
                c3.font = Font(name=font_family, size=11, bold=True, color='721C24')
                c3.fill = fail_fill

            short_error = (tc['error'][:500] + '...') if len(tc['error']) > 500 else tc['error']
            c4 = ws3.cell(row=idx, column=4, value=short_error)
            
            screenshot_att = None
            logs_att = None
            context_att = None
            for att in tc.get('attachments', []):
                t = att.get('type')
                if t == 'screenshot':
                    screenshot_att = att
                elif t == 'logs':
                    logs_att = att
                elif t == 'context':
                    context_att = att

            c5 = ws3.cell(row=idx, column=5)
            c5.alignment = Alignment(horizontal='center')
            if screenshot_att:
                c5.value = "View Screenshot"
                c5.hyperlink = screenshot_att['path']
                c5.font = link_font
            else:
                c5.value = "—"
                c5.font = regular_font

            c6 = ws3.cell(row=idx, column=6)
            c6.alignment = Alignment(horizontal='center')
            if logs_att:
                c6.value = "View Logs"
                c6.hyperlink = logs_att['path']
                c6.font = link_font
            else:
                c6.value = "—"
                c6.font = regular_font

            c7 = ws3.cell(row=idx, column=7)
            c7.alignment = Alignment(horizontal='center')
            if context_att:
                c7.value = "View Context"
                c7.hyperlink = context_att['path']
                c7.font = link_font
            else:
                c7.value = "—"
                c7.font = regular_font

            for cell in [c1, c2, c3, c4, c5, c6, c7]:
                cell.border = thin_border
                if cell not in [c3, c5, c6, c7]:
                    cell.font = regular_font
                    cell.alignment = Alignment(wrap_text=True)
                else:
                    if cell != c3:
                        cell.font = regular_font
            
            ws3.row_dimensions[idx].height = 55

    # ════════════════════════════════════════════════
    # SHEET 4: HISTORICAL TRENDS
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet(title="Trends")
    ws4.views.sheetView[0].showGridLines = True

    ws4.cell(row=2, column=2, value="Historical Build Run Trends").font = title_font
    
    headers = ["Build", "Execution Timestamp", "Git Commit", "Branch", "Duration (ms)", "Total Tests", "Passed", "Failed", "Skipped", "Pass Rate", "Flaky Rate", "Release Status"]
    for col_idx, header in enumerate(headers, 2):
        cell = ws4.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    ws4.row_dimensions[4].height = 24

    for idx, run in enumerate(history_data, 5):
        c1 = ws4.cell(row=idx, column=2, value=run["build"])
        c1.alignment = Alignment(horizontal='center')
        
        c2 = ws4.cell(row=idx, column=3, value=run["timestamp"])
        c2.alignment = Alignment(horizontal='center')
        
        c3 = ws4.cell(row=idx, column=4, value=run.get("git_commit", "N/A"))
        c3.alignment = Alignment(horizontal='center')
        
        c4 = ws4.cell(row=idx, column=5, value=run.get("branch", "N/A"))
        c4.alignment = Alignment(horizontal='center')
        
        c5 = ws4.cell(row=idx, column=6, value=run.get("duration_ms", 0))
        c5.alignment = Alignment(horizontal='right')
        
        c6 = ws4.cell(row=idx, column=7, value=run["total"])
        c6.alignment = Alignment(horizontal='center')
        
        c7 = ws4.cell(row=idx, column=8, value=run["passed"])
        c7.alignment = Alignment(horizontal='center')
        
        c8 = ws4.cell(row=idx, column=9, value=run["failed"])
        c8.alignment = Alignment(horizontal='center')
        if run["failed"] > 0:
            c8.fill = fail_fill
            c8.font = Font(name=font_family, size=11, bold=True, color='721C24')
            
        c9 = ws4.cell(row=idx, column=10, value=run.get("skipped", 0))
        c9.alignment = Alignment(horizontal='center')
        if run.get("skipped", 0) > 0:
            c9.fill = skip_fill
            c9.font = Font(name=font_family, size=11, bold=True, color='856404')
            
        c10 = ws4.cell(row=idx, column=11, value=f"{run['pass_rate']}%")
        c10.alignment = Alignment(horizontal='right')
        if run["pass_rate"] >= 90.0:
            c10.fill = pass_fill
            c10.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c10.fill = fail_fill
            c10.font = Font(name=font_family, size=11, bold=True, color='721C24')

        f_rate = run.get('flaky_rate', 0.0)
        c11 = ws4.cell(row=idx, column=12, value=f"{f_rate}%")
        c11.alignment = Alignment(horizontal='right')
        if f_rate == 0.0:
            c11.fill = pass_fill
            c11.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c11.fill = fail_fill
            c11.font = Font(name=font_family, size=11, bold=True, color='721C24')

        c12 = ws4.cell(row=idx, column=13, value=run["status"])
        c12.alignment = Alignment(horizontal='center')
        if run["status"] == 'APPROVED':
            status_col = '28A745'
        elif run["status"] == 'APPROVED WITH PERFORMANCE DEBT':
            status_col = 'FD7E14'
        elif run["status"] == 'APPROVED WITH PROTOCOL DEBT':
            status_col = '6F42C1'
        else:
            status_col = 'DC3545'
        c12.fill = PatternFill(start_color=status_col, end_color=status_col, fill_type='solid')
        c12.font = Font(name=font_family, size=11, bold=True, color='FFFFFF')

        for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11]:
            cell.border = thin_border
            if cell not in [c8, c9, c10, c11]:
                cell.font = regular_font
        c12.border = thin_border
        ws4.row_dimensions[idx].height = 20

    # ════════════════════════════════════════════════
    # SHEET 5: MAVLINK VALIDATION
    # ════════════════════════════════════════════════
    mavlink_data = kpi_data.get("mavlink", {}) if kpi_data else {}
    compliance = mavlink_data.get("compliance_score") if mavlink_data else None
    optional_coverage = mavlink_data.get("optional_coverage") if mavlink_data else None
    packet_loss = mavlink_data.get("packet_loss_pct") if mavlink_data else None
    crc_failures = mavlink_data.get("crc_failures") if mavlink_data else None

    ws5 = wb.create_sheet(title="MAVLink Validation")
    ws5.views.sheetView[0].showGridLines = True

    # Title
    ws5.cell(row=2, column=2, value="MAVLink Protocol Compliance & Message Coverage").font = title_font

    # Message coverage metrics
    ws5.cell(row=4, column=2, value="Compliance Scores").font = section_font
    ws5.cell(row=5, column=2, value="Metric").font = header_font
    ws5.cell(row=5, column=2).fill = header_fill
    ws5.cell(row=5, column=2).border = thin_border
    ws5.cell(row=5, column=3, value="Value").font = header_font
    ws5.cell(row=5, column=3).fill = header_fill
    ws5.cell(row=5, column=3).border = thin_border
    ws5.cell(row=5, column=3).alignment = Alignment(horizontal='right')

    comp_score_val = f"{compliance:.1f}%" if compliance is not None else "N/A"
    opt_cov_val = f"{optional_coverage:.1f}%" if optional_coverage is not None else "N/A"

    coverage_metrics = [
        ("Required Message Compliance Score", comp_score_val),
        ("Optional Message Coverage Score", opt_cov_val)
    ]
    for idx, (lbl, val) in enumerate(coverage_metrics, 6):
        c1 = ws5.cell(row=idx, column=2, value=lbl)
        c1.font = bold_font
        c1.fill = accent_fill
        c1.border = thin_border
        
        c2 = ws5.cell(row=idx, column=3, value=val)
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='right')
        if "Compliance" in lbl and compliance is not None:
            if compliance >= 95.0:
                c2.fill = pass_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='155724')
            elif compliance >= 90.0:
                c2.fill = skip_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='856404')
            else:
                c2.fill = fail_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        ws5.row_dimensions[idx].height = 20

    # Detailed required and optional lists
    ws5.cell(row=4, column=5, value="Protocol Message Checklists").font = section_font
    ws5.cell(row=5, column=5, value="Message Type").font = header_font
    ws5.cell(row=5, column=5).fill = header_fill
    ws5.cell(row=5, column=5).border = thin_border
    ws5.cell(row=5, column=6, value="Category").font = header_font
    ws5.cell(row=5, column=6).fill = header_fill
    ws5.cell(row=5, column=6).border = thin_border
    ws5.cell(row=5, column=7, value="Observed").font = header_font
    ws5.cell(row=5, column=7).fill = header_fill
    ws5.cell(row=5, column=7).border = thin_border
    ws5.cell(row=5, column=7).alignment = Alignment(horizontal='center')

    # Message checklists data
    required_list = [
        "HEARTBEAT", "COMMAND_LONG", "COMMAND_ACK", "PARAM_SET", "PARAM_VALUE",
        "MISSION_COUNT", "MISSION_ITEM_INT", "MISSION_ACK", "STATUSTEXT", "GLOBAL_POSITION_INT"
    ]
    optional_list = [
        "SYS_STATUS", "GPS_RAW_INT", "ATTITUDE", "VFR_HUD"
    ]
    observed_req_set = set(mavlink_data.get("observed_required", [])) if mavlink_data else set()
    observed_opt_set = set(mavlink_data.get("observed_optional", [])) if mavlink_data else set()

    chk_row = 6
    for msg in required_list:
        c1 = ws5.cell(row=chk_row, column=5, value=msg)
        c2 = ws5.cell(row=chk_row, column=6, value="Required")
        observed = msg in observed_req_set
        c3 = ws5.cell(row=chk_row, column=7, value="YES" if observed else "NO")
        
        c1.font = regular_font
        c2.font = Font(name=font_family, size=11, italic=True)
        c3.font = bold_font
        c3.alignment = Alignment(horizontal='center')
        
        for cell in [c1, c2, c3]:
            cell.border = thin_border
        
        if observed:
            c3.fill = pass_fill
            c3.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c3.fill = fail_fill
            c3.font = Font(name=font_family, size=11, bold=True, color='721C24')
        
        ws5.row_dimensions[chk_row].height = 20
        chk_row += 1

    for msg in optional_list:
        c1 = ws5.cell(row=chk_row, column=5, value=msg)
        c2 = ws5.cell(row=chk_row, column=6, value="Optional")
        observed = msg in observed_opt_set
        c3 = ws5.cell(row=chk_row, column=7, value="YES" if observed else "NO")
        
        c1.font = regular_font
        c2.font = Font(name=font_family, size=11, italic=True)
        c3.font = bold_font
        c3.alignment = Alignment(horizontal='center')
        
        for cell in [c1, c2, c3]:
            cell.border = thin_border
            
        if observed:
            c3.fill = pass_fill
            c3.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c3.fill = accent_fill
            c3.font = Font(name=font_family, size=11, bold=True, color='4A5568')
            
        ws5.row_dimensions[chk_row].height = 20
        chk_row += 1

    # ════════════════════════════════════════════════
    # SHEET 6: PACKET STATISTICS
    # ════════════════════════════════════════════════
    ws6 = wb.create_sheet(title="Packet Statistics")
    ws6.views.sheetView[0].showGridLines = True

    # Title
    ws6.cell(row=2, column=2, value="MAVLink Interception & Packet Sequence Statistics").font = title_font

    # Overview Table
    ws6.cell(row=4, column=2, value="Overview Statistics").font = section_font
    
    total_received_val = mavlink_data.get("packets_received", 0) if mavlink_data else 0
    total_lost_val = mavlink_data.get("packets_lost", 0) if mavlink_data else 0
    loss_pct_val = f"{packet_loss:.2f}%" if packet_loss is not None else "0.00%"
    crc_failures_val = crc_failures if crc_failures is not None else 0

    overview_rows = [
        ("Total Packets Received", total_received_val),
        ("Total Packets Lost (Gaps)", total_lost_val),
        ("Overall Packet Loss Rate", loss_pct_val),
        ("CRC Validation Failures", crc_failures_val)
    ]
    for idx, (lbl, val) in enumerate(overview_rows, 5):
        c1 = ws6.cell(row=idx, column=2, value=lbl)
        c1.font = bold_font
        c1.fill = accent_fill
        c1.border = thin_border
        
        c2 = ws6.cell(row=idx, column=3, value=val)
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='right')
        
        if lbl == "Overall Packet Loss Rate" and packet_loss is not None:
            if packet_loss < 1.0:
                c2.fill = pass_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='155724')
            elif packet_loss <= 3.0:
                c2.fill = skip_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='856404')
            else:
                c2.fill = fail_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        elif lbl == "CRC Validation Failures" and crc_failures_val > 0:
            c2.fill = fail_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
            
        ws6.row_dimensions[idx].height = 20

    # Directional stats table
    ws6.cell(row=11, column=2, value="Direction-Wise Sequence Analysis").font = section_font
    dir_headers = ["Direction", "Packets Received", "Sequence Gaps", "Duplicate Packets", "Out-of-Order Packets"]
    for col_idx, h in enumerate(dir_headers, 2):
        cell = ws6.cell(row=12, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left')
    ws6.row_dimensions[12].height = 20

    seq_stats = mavlink_data.get("sequence_stats", {}) if mavlink_data else {}
    
    dir_rows = [
        ("SITL → GCS (Outbound)", "SITL_TO_GCS"),
        ("GCS → SITL (Inbound)", "GCS_TO_SITL")
    ]
    for idx, (label, key) in enumerate(dir_rows, 13):
        stats_dir = seq_stats.get(key, {})
        
        c1 = ws6.cell(row=idx, column=2, value=label)
        c2 = ws6.cell(row=idx, column=3, value=stats_dir.get("received", 0))
        c3 = ws6.cell(row=idx, column=4, value=stats_dir.get("gaps", 0))
        c4 = ws6.cell(row=idx, column=5, value=stats_dir.get("duplicates", 0))
        c5 = ws6.cell(row=idx, column=6, value=stats_dir.get("out_of_order", 0))
        
        c1.font = regular_font
        for cell in [c2, c3, c4, c5]:
            cell.font = regular_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        c1.border = thin_border
        
        if stats_dir.get("gaps", 0) > 0:
            c3.fill = fail_fill
            c3.font = Font(name=font_family, size=11, bold=True, color='721C24')
        if stats_dir.get("duplicates", 0) > 0:
            c4.fill = skip_fill
        if stats_dir.get("out_of_order", 0) > 0:
            c5.fill = skip_fill
            
        ws6.row_dimensions[idx].height = 20

    # ════════════════════════════════════════════════
    # SHEET 7: TELEMETRY HEALTH & LATENCY
    # ════════════════════════════════════════════════
    ws7 = wb.create_sheet(title="Telemetry Health")
    ws7.views.sheetView[0].showGridLines = True

    # Title
    ws7.cell(row=2, column=2, value="MAVLink Telemetry Stream Frequencies, Jitter, & Latency").font = title_font

    # Telemetry Frequencies Table
    ws7.cell(row=4, column=2, value="Telemetry Stream Frequencies").font = section_font
    ws7.cell(row=5, column=2, value="Stream / Message").font = header_font
    ws7.cell(row=5, column=2).fill = header_fill
    ws7.cell(row=5, column=2).border = thin_border
    ws7.cell(row=5, column=3, value="Measured Rate (Hz)").font = header_font
    ws7.cell(row=5, column=3).fill = header_fill
    ws7.cell(row=5, column=3).border = thin_border
    ws7.cell(row=5, column=3).alignment = Alignment(horizontal='right')
    ws7.cell(row=5, column=4, value="Target Min (Hz)").font = header_font
    ws7.cell(row=5, column=4).fill = header_fill
    ws7.cell(row=5, column=4).border = thin_border
    ws7.cell(row=5, column=4).alignment = Alignment(horizontal='center')
    ws7.cell(row=5, column=5, value="Status").font = header_font
    ws7.cell(row=5, column=5).fill = header_fill
    ws7.cell(row=5, column=5).border = thin_border
    ws7.cell(row=5, column=5).alignment = Alignment(horizontal='center')
    ws7.row_dimensions[5].height = 20

    hz_heartbeat = mavlink_data.get("heartbeat_hz", 0.0) if mavlink_data else 0.0
    hz_gps = mavlink_data.get("gps_hz", 0.0) if mavlink_data else 0.0
    hz_sys = mavlink_data.get("sys_status_hz", 0.0) if mavlink_data else 0.0
    hz_attitude = mavlink_data.get("attitude_hz", 0.0) if mavlink_data else 0.0

    freqs_rows = [
        ("HEARTBEAT", hz_heartbeat, 1.0),
        ("GLOBAL_POSITION_INT", hz_gps, 5.0),
        ("SYS_STATUS", hz_sys, 1.0),
        ("ATTITUDE", hz_attitude, 1.0)
    ]
    for idx, (lbl, val, target_min) in enumerate(freqs_rows, 6):
        c1 = ws7.cell(row=idx, column=2, value=lbl)
        c2 = ws7.cell(row=idx, column=3, value=f"{val:.2f} Hz")
        c3 = ws7.cell(row=idx, column=4, value=f"{target_min:.1f} Hz")
        
        status_str = "PASS" if val >= target_min else "FAIL"
        c4 = ws7.cell(row=idx, column=5, value=status_str)
        
        c1.font = regular_font
        c2.font = regular_font
        c3.font = regular_font
        c4.font = bold_font
        
        c2.alignment = Alignment(horizontal='right')
        c3.alignment = Alignment(horizontal='center')
        c4.alignment = Alignment(horizontal='center')
        
        for cell in [c1, c2, c3, c4]:
            cell.border = thin_border
            
        if status_str == "PASS":
            c4.fill = pass_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c4.fill = fail_fill
            c4.font = Font(name=font_family, size=11, bold=True, color='721C24')
            
        ws7.row_dimensions[idx].height = 20

    # Latencies Table
    ws7.cell(row=12, column=2, value="Command Latencies & Discovery Details").font = section_font
    
    jitter_val = f"{mavlink_data.get('jitter_ms', 0.0):.1f} ms" if mavlink_data else "—"
    arm_lat_val = f"{mavlink_data.get('arm_latency_ms', 0.0):.1f} ms" if mavlink_data and mavlink_data.get('arm_latency_ms') else "—"
    mode_lat_val = f"{mavlink_data.get('mode_latency_ms', 0.0):.1f} ms" if mavlink_data and mavlink_data.get('mode_latency_ms') else "—"
    sysid_val = mavlink_data.get("last_observed_sysid", "—") if mavlink_data else "—"
    compid_val = mavlink_data.get("last_observed_compid", "—") if mavlink_data else "—"
    gcs_end_val = mavlink_data.get("latest_gcs_endpoint", "—") if mavlink_data else "—"
    dl_status_val = "COMPLETED" if mavlink_data and mavlink_data.get("download_completed") else "FAILED/INCOMPLETE"
    dl_count_val = mavlink_data.get("download_item_count", 0) if mavlink_data else 0

    latency_rows = [
        ("Heartbeat Jitter (Std Dev)", jitter_val),
        ("Arm Command-ACK Latency", arm_lat_val),
        ("Mode Switch Transition Latency", mode_lat_val),
        ("Discovered System ID", sysid_val),
        ("Discovered Component ID", compid_val),
        ("Active GCS Endpoint (UDP)", gcs_end_val),
        ("Mission Retrieval Handshake", dl_status_val),
        ("Retrieved Waypoint Count", dl_count_val)
    ]
    for idx, (lbl, val) in enumerate(latency_rows, 13):
        c1 = ws7.cell(row=idx, column=2, value=lbl)
        c1.font = bold_font
        c1.fill = accent_fill
        c1.border = thin_border
        
        c2 = ws7.cell(row=idx, column=3, value=val)
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = Alignment(horizontal='right')
        
        if lbl == "Mission Retrieval Handshake":
            if val == "COMPLETED":
                c2.fill = pass_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='155724')
            else:
                c2.fill = fail_fill
                c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
        elif lbl == "Arm Command-ACK Latency" and mavlink_data and mavlink_data.get('arm_latency_ms'):
            lat = mavlink_data.get('arm_latency_ms')
            if lat < 500:
                c2.fill = pass_fill
            else:
                c2.fill = fail_fill
        elif lbl == "Mode Switch Transition Latency" and mavlink_data and mavlink_data.get('mode_latency_ms'):
            lat = mavlink_data.get('mode_latency_ms')
            if lat < 1000:
                c2.fill = pass_fill
            else:
                c2.fill = fail_fill
                
        ws7.row_dimensions[idx].height = 20

    # ════════════════════════════════════════════════
    # SHEET 8: MULTI-VEHICLE VALIDATION
    # ════════════════════════════════════════════════
    ws8 = wb.create_sheet(title="Multi-Vehicle Validation")
    ws8.views.sheetView[0].showGridLines = True

    # Title
    ws8.cell(row=2, column=2, value="Multi-Vehicle Fleet Routing & Isolation Validation").font = title_font

    ws8.cell(row=4, column=2, value="Fleet Validation Metrics").font = section_font
    ws8.cell(row=5, column=2, value="Validation Metric").font = header_font
    ws8.cell(row=5, column=2).fill = header_fill
    ws8.cell(row=5, column=2).border = thin_border
    ws8.cell(row=5, column=3, value="Result").font = header_font
    ws8.cell(row=5, column=3).fill = header_fill
    ws8.cell(row=5, column=3).border = thin_border
    ws8.cell(row=5, column=3).alignment = Alignment(horizontal='center')
    ws8.row_dimensions[5].height = 20

    mv_data = kpi_data.get("multi_vehicle", {}) if kpi_data else {}
    
    disc_status = "PASS" if mv_data.get("discovery_success_rate", 0) == 100 else "FAIL"
    veh_iso = mv_data.get("vehicle_isolation", "FAIL")
    param_iso = mv_data.get("parameter_isolation", "FAIL")
    mission_iso = mv_data.get("mission_isolation", "FAIL")
    cmd_iso = mv_data.get("command_isolation", "FAIL")
    disc_rec = mv_data.get("disconnect_recovery", "FAIL")
    rec_rec = mv_data.get("reconnect_recovery", "FAIL")

    mv_rows = [
        ("Vehicle Discovery", disc_status),
        ("Vehicle Isolation", veh_iso),
        ("Parameter Isolation", param_iso),
        ("Mission Isolation", mission_iso),
        ("Command Isolation", cmd_iso),
        ("Disconnect Recovery", disc_rec),
        ("Reconnect Recovery", rec_rec)
    ]

    for idx, (lbl, status_str) in enumerate(mv_rows, 6):
        c1 = ws8.cell(row=idx, column=2, value=lbl)
        c2 = ws8.cell(row=idx, column=3, value=status_str)
        
        c1.font = regular_font
        c2.font = bold_font
        c2.alignment = Alignment(horizontal='center')
        
        for cell in [c1, c2]:
            cell.border = thin_border
            
        if status_str == "PASS":
            c2.fill = pass_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='155724')
        else:
            c2.fill = fail_fill
            c2.font = Font(name=font_family, size=11, bold=True, color='721C24')
            
        ws8.row_dimensions[idx].height = 20

    # ════════════════════════════════════════════════
    # AUTO-FIT COLUMN WIDTHS (ALL SHEETS)
    # ════════════════════════════════════════════════
    for sheet in [ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            if sheet == ws3 and col_letter == 'D':
                sheet.column_dimensions[col_letter].width = 60
                continue
            if sheet == ws3 and col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 35
                continue

            for cell in col:
                if cell.row in [2, 3] and sheet != ws2:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    try:
        wb.save(EXCEL_PATH)
        print(f"[Report] Integration test Excel report generated successfully at: {EXCEL_PATH}")
    except (PermissionError, IOError) as e:
        print(f"\n[Warning] Permission denied writing to {EXCEL_PATH}. Please ensure the file is closed in Excel.")
        alternative_path = EXCEL_PATH.replace('.xlsx', '_locked.xlsx')
        try:
            wb.save(alternative_path)
            print(f"[Report] Saved alternative report copy to: {alternative_path}")
        except Exception as alt_err:
            print(f"Error saving alternative copy: {alt_err}")

if __name__ == '__main__':
    main()
